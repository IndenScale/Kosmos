"""
Office文档解析器
文件: office_parsers.py
创建时间: 2024-12-19
描述: 实现Word、Excel、PowerPoint等Office文档的解析器，集成智能文本分割器
"""

import os
import uuid
import subprocess
from typing import List, Dict, Any, Optional, Tuple
from pathlib import Path
import zipfile
import xml.etree.ElementTree as ET
import shutil

from app.parsers.base_parser import DocumentParser, ParsedFragment
from app.parsers.text_parsers import MarkdownParser
from app.parsers.parser_utils import url_to_local_path, PageRangeUtils
from app.schemas.fragment import FragmentType
from app.utils.text_splitter import TextSplitter


class DocxParser(DocumentParser):
    """Word文档解析器"""
    
    def can_parse(self, file_path: str, mime_type: str) -> bool:
        """判断是否可以解析该文件"""
        local_path = url_to_local_path(file_path)
        return (mime_type == 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' or
                Path(local_path).suffix.lower() == '.docx')
    
    def parse(self, file_path: str) -> List[ParsedFragment]:
        """解析Word文档 - 转换为PDF后使用PDF解析器处理"""
        try:
            # 首先尝试转换为PDF并使用PDF解析器
            try:
                return self._parse_via_pdf_conversion(file_path)
            except Exception as pdf_error:
                self.logger.warning(f"PDF转换解析失败，回退到文本解析: {pdf_error}")
                # 回退到原始的文本解析方法
                return self._parse_text_only(file_path)
            
        except Exception as e:
            self.logger.error(f"Word文档解析失败: {file_path}, 错误: {e}")
            return self._create_error_fragment(file_path, str(e))
    
    def _parse_via_pdf_conversion(self, file_path: str) -> List[ParsedFragment]:
        """通过转换为PDF来解析DOCX文件"""
        # 创建临时目录
        temp_id = str(uuid.uuid4())
        temp_dir = Path("data/temp/docx_conversion") / temp_id
        temp_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            # 转换DOCX为PDF
            pdf_path = self._convert_to_pdf(file_path, temp_dir)
            
            # 使用PDF解析器解析转换后的PDF
            from app.parsers.pdf_parser import PdfParser
            pdf_parser = PdfParser(self.db, self.kb_id)
            
            # 解析PDF
            fragments = pdf_parser.parse(pdf_path)
            
            # 更新meta_info以反映原始文件类型
            for fragment in fragments:
                if fragment.meta_info:
                    fragment.meta_info['original_source_type'] = 'docx'
                    fragment.meta_info['original_source_file'] = Path(file_path).name
                    fragment.meta_info['conversion_method'] = 'docx_to_pdf'
            
            self.logger.info(f"DOCX通过PDF转换解析成功: {file_path}, 生成{len(fragments)}个fragments")
            return fragments
            
        finally:
            # 清理临时目录
            if temp_dir.exists():
                shutil.rmtree(temp_dir)
    
    def _convert_to_pdf(self, docx_path: str, output_dir: Path) -> str:
        """将DOCX转换为PDF"""
        local_path = url_to_local_path(docx_path)
        docx_file = Path(local_path)
        pdf_filename = f"{docx_file.stem}.pdf"
        pdf_path = output_dir / pdf_filename

        # 按优先级尝试不同的转换方法
        conversion_methods = [
            ("LibreOffice", self._convert_with_libreoffice),
            ("docx2pdf", self._convert_with_docx2pdf),
            ("Microsoft Word COM", self._convert_with_word_com)
        ]

        for method_name, method_func in conversion_methods:
            try:
                self.logger.info(f"尝试使用 {method_name} 转换 DOCX 到 PDF...")
                method_func(local_path, str(pdf_path))
                
                if pdf_path.exists():
                    self.logger.info(f"DOCX转PDF成功 ({method_name}): {local_path} -> {pdf_path}")
                    return str(pdf_path)
                else:
                    self.logger.warning(f"{method_name} 转换完成但PDF文件不存在")
                    
            except Exception as e:
                self.logger.warning(f"{method_name} 转换失败: {str(e)}")
                continue

        # 如果所有方法都失败，抛出异常
        raise Exception("所有DOCX转PDF方法都失败，请检查系统环境或安装相关依赖")
    
    def _convert_with_libreoffice(self, docx_path: str, pdf_path: str):
        """使用LibreOffice转换"""
        if not self._check_libreoffice_available():
            raise Exception("LibreOffice不可用")

        output_dir = Path(pdf_path).parent
        cmd = [
            'libreoffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', str(output_dir),
            str(docx_path)
        ]

        result = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            check=True,
            timeout=60
        )

        if result.returncode != 0:
            raise Exception(f"LibreOffice转换失败: {result.stderr}")

    def _convert_with_docx2pdf(self, docx_path: str, pdf_path: str):
        """使用docx2pdf库转换（仅Windows）"""
        try:
            from docx2pdf import convert
            convert(docx_path, pdf_path)
        except ImportError:
            raise Exception("docx2pdf库未安装，请运行: pip install docx2pdf")
        except Exception as e:
            raise Exception(f"docx2pdf转换失败: {str(e)}")

    def _convert_with_word_com(self, docx_path: str, pdf_path: str):
        """使用Microsoft Word COM接口转换（仅Windows）"""
        import platform
        if platform.system() != "Windows":
            raise Exception("Word COM接口仅在Windows上可用")

        try:
            import comtypes.client
            
            # 启动Word应用程序
            word = comtypes.client.CreateObject('Word.Application')
            word.Visible = False
            
            try:
                # 打开DOCX文档
                doc = word.Documents.Open(os.path.abspath(docx_path))
                
                # 导出为PDF
                doc.ExportAsFixedFormat(
                    OutputFileName=os.path.abspath(pdf_path),
                    ExportFormat=17,  # PDF格式
                    OpenAfterExport=False,
                    OptimizeFor=0,
                    BitmapMissingFonts=True,
                    DocStructureTags=False,
                    CreateBookmarks=False
                )
                
                doc.Close()
            finally:
                word.Quit()
                
        except ImportError:
            raise Exception("comtypes库未安装或Microsoft Word未安装")
        except Exception as e:
            raise Exception(f"Word COM转换失败: {str(e)}")

    def _check_libreoffice_available(self) -> bool:
        """检查LibreOffice是否可用"""
        try:
            subprocess.run(
                ['libreoffice', '--version'],
                capture_output=True,
                check=True,
                timeout=10
            )
            return True
        except (subprocess.CalledProcessError, FileNotFoundError, subprocess.TimeoutExpired):
            return False
    
    def _parse_text_only(self, file_path: str) -> List[ParsedFragment]:
        """原始的纯文本解析方法（作为回退方案）"""
        try:
            # 提取文本内容
            text_content = self._extract_docx_text(file_path)
            
            if not text_content.strip():
                return self._create_empty_fragment(file_path)
            
            # 使用Markdown解析器分割内容
            return self._split_text_content(text_content, file_path)
            
        except Exception as e:
            self.logger.error(f"Word文档文本解析失败: {file_path}, 错误: {e}")
            return self._create_error_fragment(file_path, str(e))
    
    def _split_text_content(self, text_content: str, file_path: str) -> List[ParsedFragment]:
        """分割文本内容为fragments"""
        if not text_content.strip():
            return []
        
        # 使用智能文本分割器
        try:
            splitter = TextSplitter(
                chunk_size=1350,
                chunk_overlap=150
            )
            
            chunks = splitter.split_text(text_content)
            fragments = []
            
            # 解析文本内容中的页面信息
            page_ranges = PageRangeUtils.extract_page_info(text_content)
            
            for i, chunk in enumerate(chunks):
                if chunk.strip():  # 跳过空白块
                    # 确定chunk的页面范围
                    page_start, page_end = PageRangeUtils.determine_chunk_pages(chunk, page_ranges, text_content)
                    
                    meta_info = self._create_meta_info(
                        content_length=len(chunk),
                        chunk_index=i,
                        total_chunks=len(chunks),
                        source_type='docx',
                        source_file=Path(file_path).name,
                        processed_by='docx_parser',
                        split_method='intelligent_splitter',
                        page_start=page_start,
                        page_end=page_end
                    )
                    
                    fragment = ParsedFragment(
                        fragment_type=FragmentType.TEXT,
                        raw_content=chunk,
                        meta_info=meta_info,
                        fragment_index=i,
                        page_start=page_start,
                        page_end=page_end
                    )
                    fragments.append(fragment)
            
            return fragments
            
        except Exception as e:
            self.logger.warning(f"智能分割失败，使用段落分割: {e}")
            # 回退到段落分割
            return self._split_text_by_paragraphs(text_content, file_path)
    
    def _split_text_by_paragraphs(self, text_content: str, file_path: str) -> List[ParsedFragment]:
        """按段落分割文本内容"""
        paragraphs = [p.strip() for p in text_content.split('\n\n') if p.strip()]
        
        fragments = []
        # 解析文本内容中的页面信息
        page_ranges = PageRangeUtils.extract_page_info(text_content)
        
        for i, paragraph in enumerate(paragraphs):
            if len(paragraph) < 50:  # 跳过太短的段落
                continue
            
            # 确定段落的页面范围
            page_start, page_end = PageRangeUtils.determine_chunk_pages(paragraph, page_ranges, text_content)
            
            meta_info = self._create_meta_info(
                content_length=len(paragraph),
                paragraph_index=i,
                total_paragraphs=len(paragraphs),
                source_type='docx',
                source_file=Path(file_path).name,
                processed_by='docx_parser',
                split_method='paragraph_fallback',
                page_start=page_start,
                page_end=page_end
            )
            
            fragment = ParsedFragment(
                fragment_type=FragmentType.TEXT,
                raw_content=paragraph,
                meta_info=meta_info,
                fragment_index=i,
                page_start=page_start,
                page_end=page_end
            )
            fragments.append(fragment)
        
        return fragments
    
    def _create_empty_fragment(self, file_path: str) -> List[ParsedFragment]:
        """创建空Fragment"""
        meta_info = self._create_meta_info(
            content_length=0,
            source_file=Path(file_path).name,
            source_type='docx',
            processed_by='docx_parser'
        )
        
        fragment = ParsedFragment(
            fragment_type=FragmentType.TEXT,
            raw_content=f"文档为空: {Path(file_path).name}",
            meta_info=meta_info,
            fragment_index=0
        )
        
        return [fragment]
    
    def _create_error_fragment(self, file_path: str, error_message: str) -> List[ParsedFragment]:
        """创建错误Fragment"""
        meta_info = self._create_meta_info(
            error_type="parsing_error",
            error_message=error_message,
            source_file=Path(file_path).name,
            source_type='docx',
            processed_by='docx_parser'
        )
        
        fragment = ParsedFragment(
            fragment_type=FragmentType.TEXT,
            raw_content=f"文档解析失败: {Path(file_path).name}\n错误信息: {error_message}",
            meta_info=meta_info,
            fragment_index=0
        )
        
        return [fragment]
    
    def _extract_docx_text(self, file_path: str) -> str:
        """从DOCX文件提取文本"""
        try:
            # 转换URL为本地路径
            local_path = url_to_local_path(file_path)
            with zipfile.ZipFile(local_path, 'r') as docx:
                # 读取主文档内容
                document_xml = docx.read('word/document.xml')
                root = ET.fromstring(document_xml)
                
                # 定义命名空间
                namespaces = {
                    'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
                }
                
                # 提取所有文本
                text_parts = []
                
                # 提取段落
                for para in root.findall('.//w:p', namespaces):
                    para_text = self._extract_paragraph_text(para, namespaces)
                    if para_text.strip():
                        text_parts.append(para_text)
                
                # 提取表格
                for table in root.findall('.//w:tbl', namespaces):
                    table_text = self._extract_table_text(table, namespaces)
                    if table_text.strip():
                        text_parts.append(table_text)
                
                return '\n\n'.join(text_parts)
                
        except Exception as e:
            self.logger.error(f"DOCX文本提取失败: {file_path}, 错误: {e}")
            raise
    
    def _extract_paragraph_text(self, para, namespaces: Dict[str, str]) -> str:
        """提取段落文本"""
        text_parts = []
        
        for run in para.findall('.//w:r', namespaces):
            for text_elem in run.findall('.//w:t', namespaces):
                if text_elem.text:
                    text_parts.append(text_elem.text)
        
        return ''.join(text_parts)
    
    def _extract_table_text(self, table, namespaces: Dict[str, str]) -> str:
        """提取表格文本"""
        table_rows = []
        
        for row in table.findall('.//w:tr', namespaces):
            row_cells = []
            for cell in row.findall('.//w:tc', namespaces):
                cell_text = []
                for para in cell.findall('.//w:p', namespaces):
                    para_text = self._extract_paragraph_text(para, namespaces)
                    if para_text.strip():
                        cell_text.append(para_text)
                row_cells.append(' '.join(cell_text))
            
            if any(cell.strip() for cell in row_cells):
                table_rows.append(' | '.join(row_cells))
        
        if table_rows:
            return '\n**表格:**\n' + '\n'.join(table_rows) + '\n'
        
        return ''


class XlsxParser(DocumentParser):
    """Excel文档解析器"""
    
    def can_parse(self, file_path: str, mime_type: str) -> bool:
        """判断是否可以解析该文件"""
        local_path = url_to_local_path(file_path)
        return (mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' or
                Path(local_path).suffix.lower() == '.xlsx')
    
    def parse(self, file_path: str) -> List[ParsedFragment]:
        """解析Excel文档"""
        try:
            # 提取工作表内容
            sheets_content = self._extract_xlsx_content(file_path)
            
            if not sheets_content:
                return self._create_empty_fragment(file_path)
            
            fragments = []
            for i, (sheet_name, content) in enumerate(sheets_content.items()):
                if content.strip():
                    sheet_fragments = self._create_sheet_fragments(sheet_name, content, i)
                    if isinstance(sheet_fragments, list):
                        fragments.extend(sheet_fragments)
                    else:
                        fragments.append(sheet_fragments)
            
            return fragments if fragments else self._create_empty_fragment(file_path)
            
        except Exception as e:
            self.logger.error(f"Excel文档解析失败: {file_path}, 错误: {e}")
            return self._create_error_fragment(file_path, str(e))
    
    def _extract_xlsx_content(self, file_path: str) -> Dict[str, str]:
        """从XLSX文件提取内容"""
        try:
            import openpyxl
            
            # 转换URL为本地路径
            local_path = url_to_local_path(file_path)
            workbook = openpyxl.load_workbook(local_path, data_only=True)
            sheets_content = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 提取工作表数据
                rows_data = []
                for row in sheet.iter_rows(values_only=True):
                    # 过滤空行
                    if any(cell is not None and str(cell).strip() for cell in row):
                        row_data = [str(cell) if cell is not None else '' for cell in row]
                        rows_data.append(' | '.join(row_data))
                
                if rows_data:
                    content = f"**工作表: {sheet_name}**\n\n" + '\n'.join(rows_data)
                    sheets_content[sheet_name] = content
            
            workbook.close()
            return sheets_content
            
        except ImportError:
            # 如果没有openpyxl，使用基础XML解析
            return self._extract_xlsx_content_xml(file_path)
        except Exception as e:
            self.logger.error(f"XLSX内容提取失败: {file_path}, 错误: {e}")
            raise
    
    def _extract_xlsx_content_xml(self, file_path: str) -> Dict[str, str]:
        """使用XML解析XLSX文件（备用方案）"""
        try:
            # 转换URL为本地路径
            local_path = url_to_local_path(file_path)
            with zipfile.ZipFile(local_path, 'r') as xlsx:
                # 读取共享字符串
                shared_strings = self._read_shared_strings(xlsx)
                
                # 读取工作表
                sheets_content = {}
                
                # 获取工作簿信息
                workbook_xml = xlsx.read('xl/workbook.xml')
                workbook_root = ET.fromstring(workbook_xml)
                
                namespaces = {
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }
                
                # 遍历工作表
                for sheet in workbook_root.findall('.//sheet', namespaces):
                    sheet_name = sheet.get('name', 'Unknown')
                    sheet_id = sheet.get('sheetId', '1')
                    
                    try:
                        sheet_xml = xlsx.read(f'xl/worksheets/sheet{sheet_id}.xml')
                        sheet_content = self._parse_sheet_xml(sheet_xml, shared_strings)
                        
                        if sheet_content.strip():
                            sheets_content[sheet_name] = f"**工作表: {sheet_name}**\n\n{sheet_content}"
                    except KeyError:
                        continue
                
                return sheets_content
                
        except Exception as e:
            self.logger.error(f"XLSX XML解析失败: {file_path}, 错误: {e}")
            return {}
    
    def _read_shared_strings(self, xlsx: zipfile.ZipFile) -> List[str]:
        """读取共享字符串表"""
        try:
            shared_strings_xml = xlsx.read('xl/sharedStrings.xml')
            root = ET.fromstring(shared_strings_xml)
            
            strings = []
            for si in root.findall('.//si'):
                text_parts = []
                for t in si.findall('.//t'):
                    if t.text:
                        text_parts.append(t.text)
                strings.append(''.join(text_parts))
            
            return strings
        except KeyError:
            return []
    
    def _parse_sheet_xml(self, sheet_xml: bytes, shared_strings: List[str]) -> str:
        """解析工作表XML"""
        root = ET.fromstring(sheet_xml)
        
        rows_data = []
        for row in root.findall('.//row'):
            row_data = []
            for cell in row.findall('.//c'):
                cell_value = self._get_cell_value(cell, shared_strings)
                if cell_value:
                    row_data.append(cell_value)
            
            if row_data:
                rows_data.append(' | '.join(row_data))
        
        return '\n'.join(rows_data)
    
    def _get_cell_value(self, cell, shared_strings: List[str]) -> str:
        """获取单元格值"""
        cell_type = cell.get('t', '')
        value_elem = cell.find('.//v')
        
        if value_elem is None or value_elem.text is None:
            return ''
        
        if cell_type == 's':  # 共享字符串
            try:
                index = int(value_elem.text)
                return shared_strings[index] if index < len(shared_strings) else ''
            except (ValueError, IndexError):
                return ''
        else:
            return value_elem.text
    
    def _create_sheet_fragments(self, sheet_name: str, content: str, index: int):
        """创建工作表Fragment"""
        # 如果内容过长，使用智能分割器
        if len(content) > 1500:
            try:
                splitter = TextSplitter(
                    chunk_size=1350,
                    chunk_overlap=150
                )
                
                chunks = splitter.split_text(content)
                fragments = []
                
                # 解析文本内容中的页面信息
                page_ranges = PageRangeUtils.extract_page_info(content)
                
                for i, chunk in enumerate(chunks):
                    # 确定chunk的页面范围
                    page_start, page_end = PageRangeUtils.determine_chunk_pages(chunk, page_ranges, content)
                    
                    meta_info = self._create_meta_info(
                        sheet_name=sheet_name,
                        content_length=len(chunk),
                        sheet_index=index,
                        chunk_index=i,
                        total_chunks=len(chunks),
                        source_type='xlsx',
                        processed_by='xlsx_parser',
                        split_method='intelligent_splitter',
                        page_start=page_start,
                        page_end=page_end
                    )
                    
                    fragment = ParsedFragment(
                        fragment_type=FragmentType.TEXT,
                        raw_content=chunk,
                        meta_info=meta_info,
                        fragment_index=index * 1000 + i,  # 确保唯一性
                        page_start=page_start,
                        page_end=page_end
                    )
                    fragments.append(fragment)
                
                return fragments
                
            except Exception as e:
                self.logger.warning(f"智能分割失败，使用原始内容: {e}")
        
        # 原始方法或回退方案
        meta_info = self._create_meta_info(
            sheet_name=sheet_name,
            content_length=len(content),
            sheet_index=index,
            source_type='xlsx',
            processed_by='xlsx_parser',
            split_method='original' if len(content) <= 1500 else 'fallback'
        )
        
        return ParsedFragment(
            fragment_type=FragmentType.TEXT,
            raw_content=content,
            meta_info=meta_info,
            fragment_index=index
        )
    
    def _create_empty_fragment(self, file_path: str) -> List[ParsedFragment]:
        """创建空Fragment"""
        meta_info = self._create_meta_info(
            content_length=0,
            source_file=Path(file_path).name,
            source_type='xlsx',
            processed_by='xlsx_parser'
        )
        
        fragment = ParsedFragment(
            fragment_type=FragmentType.TEXT,
            raw_content=f"Excel文档为空: {Path(file_path).name}",
            meta_info=meta_info,
            fragment_index=0
        )
        
        return [fragment]
    
    def _create_error_fragment(self, file_path: str, error_message: str) -> List[ParsedFragment]:
        """创建错误Fragment"""
        meta_info = self._create_meta_info(
            error_type="parsing_error",
            error_message=error_message,
            source_file=Path(file_path).name,
            source_type='xlsx',
            processed_by='xlsx_parser'
        )
        
        fragment = ParsedFragment(
            fragment_type=FragmentType.TEXT,
            raw_content=f"Excel文档解析失败: {Path(file_path).name}\n错误信息: {error_message}",
            meta_info=meta_info,
            fragment_index=0
        )
        
        return [fragment]


class PptxParser(DocumentParser):
    """PowerPoint文档解析器"""
    
    def can_parse(self, file_path: str, mime_type: str) -> bool:
        """判断是否可以解析该文件"""
        local_path = url_to_local_path(file_path)
        return (mime_type == 'application/vnd.openxmlformats-officedocument.presentationml.presentation' or
                Path(local_path).suffix.lower() == '.pptx')
    
    def parse(self, file_path: str) -> List[ParsedFragment]:
        """解析PowerPoint文档"""
        try:
            # 提取幻灯片内容
            slides_content = self._extract_pptx_content(file_path)
            
            if not slides_content:
                return self._create_empty_fragment(file_path)
            
            fragments = []
            for i, content in enumerate(slides_content):
                if content.strip():
                    fragment = self._create_slide_fragment(content, i + 1, i)
                    fragments.append(fragment)
            
            return fragments if fragments else self._create_empty_fragment(file_path)
            
        except Exception as e:
            self.logger.error(f"PowerPoint文档解析失败: {file_path}, 错误: {e}")
            return self._create_error_fragment(file_path, str(e))
    
    def _extract_pptx_content(self, file_path: str) -> List[str]:
        """从PPTX文件提取内容"""
        try:
            # 转换URL为本地路径
            local_path = url_to_local_path(file_path)
            with zipfile.ZipFile(local_path, 'r') as pptx:
                slides_content = []
                
                # 获取幻灯片列表
                slide_files = [f for f in pptx.namelist() if f.startswith('ppt/slides/slide') and f.endswith('.xml')]
                slide_files.sort()  # 确保顺序
                
                for slide_file in slide_files:
                    try:
                        slide_xml = pptx.read(slide_file)
                        slide_content = self._parse_slide_xml(slide_xml)
                        slides_content.append(slide_content)
                    except Exception as e:
                        self.logger.error(f"幻灯片解析失败: {slide_file}, 错误: {e}")
                        slides_content.append(f"[幻灯片解析失败: {slide_file}]")
                
                return slides_content
                
        except Exception as e:
            self.logger.error(f"PPTX内容提取失败: {file_path}, 错误: {e}")
            raise
    
    def _parse_slide_xml(self, slide_xml: bytes) -> str:
        """解析幻灯片XML"""
        root = ET.fromstring(slide_xml)
        
        # 定义命名空间
        namespaces = {
            'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
            'p': 'http://schemas.openxmlformats.org/presentationml/2006/main'
        }
        
        text_parts = []
        
        # 提取所有文本
        for text_elem in root.findall('.//a:t', namespaces):
            if text_elem.text and text_elem.text.strip():
                text_parts.append(text_elem.text.strip())
        
        return '\n'.join(text_parts) if text_parts else ''
    
    def _create_slide_fragment(self, content: str, slide_num: int, index: int) -> ParsedFragment:
        """创建幻灯片Fragment"""
        formatted_content = f"**幻灯片 {slide_num}**\n\n{content}"
        
        meta_info = self._create_meta_info(
            slide_number=slide_num,
            content_length=len(content),
            slide_index=index,
            source_type='pptx',
            processed_by='pptx_parser'
        )
        
        return ParsedFragment(
            fragment_type=FragmentType.TEXT,
            raw_content=formatted_content,
            meta_info=meta_info,
            fragment_index=index
        )
    
    def _create_empty_fragment(self, file_path: str) -> List[ParsedFragment]:
        """创建空Fragment"""
        meta_info = self._create_meta_info(
            content_length=0,
            source_file=Path(file_path).name,
            source_type='pptx',
            processed_by='pptx_parser'
        )
        
        fragment = ParsedFragment(
            fragment_type=FragmentType.TEXT,
            raw_content=f"PowerPoint文档为空: {Path(file_path).name}",
            meta_info=meta_info,
            fragment_index=0
        )
        
        return [fragment]
    
    def _create_error_fragment(self, file_path: str, error_message: str) -> List[ParsedFragment]:
        """创建错误Fragment"""
        meta_info = self._create_meta_info(
            error_type="parsing_error",
            error_message=error_message,
            source_file=Path(file_path).name,
            source_type='pptx',
            processed_by='pptx_parser'
        )
        
        fragment = ParsedFragment(
            fragment_type=FragmentType.TEXT,
            raw_content=f"PowerPoint文档解析失败: {Path(file_path).name}\n错误信息: {error_message}",
            meta_info=meta_info,
            fragment_index=0
        )
        
        return [fragment]